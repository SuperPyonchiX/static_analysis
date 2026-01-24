"""分類結果のExcel出力モジュール。"""

from typing import Dict, List, Optional
from pathlib import Path
import shutil
import logging

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from ..models.classification import ClassificationResult, ClassificationType

logger = logging.getLogger(__name__)


class ExcelWriter:
    """分類結果をExcelファイルに書き込む。"""

    # 各分類タイプの色（RGB hex、#なし）
    CLASSIFICATION_COLORS: Dict[ClassificationType, str] = {
        ClassificationType.FALSE_POSITIVE: "C6EFCE",  # 緑 - 問題なし
        ClassificationType.DEVIATION: "FFEB9C",       # 黄 - レビュー必要
        ClassificationType.FIX_REQUIRED: "FFC7CE",    # 赤 - 修正必要
        ClassificationType.UNDETERMINED: "D9D9D9",    # 灰 - 判定不可
    }

    # 出力先列名のマッピング（優先順位順）
    OUTPUT_COLUMNS: Dict[str, List[str]] = {
        "classification": ["対応方針", "分類", "Classification"],
        "reason": ["判断理由", "分類理由", "理由", "Reason"],
    }

    # 日本語列ヘッダー（フォールバック用）
    RESULT_HEADERS = ["分類", "分類理由", "確信度", "判定フェーズ"]

    def __init__(
        self,
        input_file: str,
        output_file: str,
        sheet_name: Optional[str] = None
    ):
        """Excelライターを初期化する。

        Args:
            input_file: 入力Excelファイルのパス
            output_file: 出力Excelファイルのパス
            sheet_name: 変更するシート名（Noneの場合はアクティブシート）
        """
        self.input_file = Path(input_file)
        self.output_file = Path(output_file)
        self.sheet_name = sheet_name

    def write_results(
        self,
        results: Dict[str, ClassificationResult],
        finding_id_to_row: Dict[str, int]
    ) -> None:
        """分類結果をExcelファイルに書き込む。

        既存の「対応方針」「判断理由」列に書き込む。
        列が存在しない場合は末尾に新規作成する。

        Args:
            results: 指摘IDから分類結果へのマッピング
            finding_id_to_row: 指摘IDからExcel行番号へのマッピング
        """
        # 入力ファイルを出力先にコピー
        shutil.copy(self.input_file, self.output_file)

        # ワークブックを開く
        wb = load_workbook(self.output_file)
        ws = wb.active if self.sheet_name is None else wb[self.sheet_name]

        # 出力先列を検索または作成
        classification_col = self._find_or_create_column(ws, "classification")
        reason_col = self._find_or_create_column(ws, "reason")

        logger.debug(f"分類列: {classification_col}, 理由列: {reason_col}")

        # 各指摘の結果を書き込む
        for finding_id, result in results.items():
            if finding_id not in finding_id_to_row:
                logger.warning(f"Finding {finding_id} not found in row mapping")
                continue

            row_num = finding_id_to_row[finding_id]
            self._write_result_to_existing_columns(
                ws, row_num, classification_col, reason_col, result
            )

        # 列幅を調整（新規作成した列のみ）
        self._adjust_output_column_widths(ws, classification_col, reason_col)

        # ワークブックを保存
        wb.save(self.output_file)
        logger.info(f"Results written to {self.output_file}")

    def _find_or_create_column(self, ws, column_type: str) -> int:
        """出力先列を検索し、なければ作成する。

        Args:
            ws: ワークシートオブジェクト
            column_type: "classification" または "reason"

        Returns:
            列番号（1始まり）
        """
        column_names = self.OUTPUT_COLUMNS[column_type]

        # ヘッダー行から列名を検索
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value in column_names:
                logger.debug(f"列 '{column_type}' を発見: 列{col} ('{header_value}')")
                return col

        # 見つからない場合は末尾に追加
        new_col = ws.max_column + 1
        default_name = column_names[0]  # デフォルト名を使用
        ws.cell(row=1, column=new_col).value = default_name
        logger.info(f"列 '{default_name}' を新規作成: 列{new_col}")
        return new_col

    def _write_result_to_existing_columns(
        self,
        ws,
        row_num: int,
        classification_col: int,
        reason_col: int,
        result: ClassificationResult
    ) -> None:
        """既存列に結果を書き込む。

        Args:
            ws: ワークシートオブジェクト
            row_num: 書き込む行番号
            classification_col: 分類列の列番号
            reason_col: 理由列の列番号
            result: 書き込む分類結果
        """
        # 分類結果（対応方針）
        cell_classification = ws.cell(row=row_num, column=classification_col)
        cell_classification.value = result.classification.value
        cell_classification.fill = PatternFill(
            start_color=self.CLASSIFICATION_COLORS[result.classification],
            end_color=self.CLASSIFICATION_COLORS[result.classification],
            fill_type="solid"
        )
        cell_classification.alignment = Alignment(horizontal="center")

        # 判断理由（確信度とフェーズ情報を含む）
        cell_reason = ws.cell(row=row_num, column=reason_col)
        cell_reason.value = f"{result.reason} (確信度: {result.confidence:.0%}, Phase: {result.phase})"
        cell_reason.alignment = Alignment(wrap_text=True, vertical="top")

    def _adjust_output_column_widths(
        self,
        ws,
        classification_col: int,
        reason_col: int
    ) -> None:
        """出力列の列幅を調整する。

        Args:
            ws: ワークシートオブジェクト
            classification_col: 分類列の列番号
            reason_col: 理由列の列番号
        """
        # 分類列
        col_letter = ws.cell(row=1, column=classification_col).column_letter
        ws.column_dimensions[col_letter].width = 12

        # 理由列
        col_letter = ws.cell(row=1, column=reason_col).column_letter
        ws.column_dimensions[col_letter].width = 60

    def _add_headers(self, ws, last_col: int) -> None:
        """結果列のヘッダーを追加する。

        Args:
            ws: ワークシートオブジェクト
            last_col: 既存の最終列インデックス
        """
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        white_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for i, header in enumerate(self.RESULT_HEADERS, 1):
            cell = ws.cell(row=1, column=last_col + i)
            cell.value = header
            cell.font = white_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border

    def _write_result_row(
        self,
        ws,
        row_num: int,
        last_col: int,
        result: ClassificationResult
    ) -> None:
        """1行分の結果を書き込む。

        Args:
            ws: ワークシートオブジェクト
            row_num: 書き込む行番号
            last_col: 既存の最終列インデックス
            result: 書き込む分類結果
        """
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 分類
        cell_classification = ws.cell(row=row_num, column=last_col + 1)
        cell_classification.value = result.classification.value
        cell_classification.fill = PatternFill(
            start_color=self.CLASSIFICATION_COLORS[result.classification],
            end_color=self.CLASSIFICATION_COLORS[result.classification],
            fill_type="solid"
        )
        cell_classification.alignment = Alignment(horizontal="center")
        cell_classification.border = thin_border

        # 理由
        cell_reason = ws.cell(row=row_num, column=last_col + 2)
        cell_reason.value = result.reason
        cell_reason.alignment = Alignment(wrap_text=True, vertical="top")
        cell_reason.border = thin_border

        # 確信度
        cell_confidence = ws.cell(row=row_num, column=last_col + 3)
        cell_confidence.value = f"{result.confidence:.0%}"
        cell_confidence.alignment = Alignment(horizontal="center")
        cell_confidence.border = thin_border

        # フェーズ
        cell_phase = ws.cell(row=row_num, column=last_col + 4)
        cell_phase.value = result.phase
        cell_phase.alignment = Alignment(horizontal="center")
        cell_phase.border = thin_border

    def _adjust_column_widths(self, ws, last_col: int) -> None:
        """結果列の列幅を調整する。

        Args:
            ws: ワークシートオブジェクト
            last_col: 既存の最終列インデックス
        """
        widths = [12, 60, 10, 12]  # 各結果列の幅

        for i, width in enumerate(widths, 1):
            col_letter = ws.cell(row=1, column=last_col + i).column_letter
            ws.column_dimensions[col_letter].width = width

    def write_summary(self, results: List[ClassificationResult]) -> None:
        """統計情報を含むサマリーシートを追加する。

        Args:
            results: 全分類結果のリスト
        """
        wb = load_workbook(self.output_file)

        # 既存のSummaryシートがあれば削除
        if "Summary" in wb.sheetnames:
            del wb["Summary"]

        # 新しいSummaryシートを作成
        ws = wb.create_sheet("Summary")

        # 統計を計算
        total = len(results)
        counts: Dict[ClassificationType, int] = {
            ClassificationType.FALSE_POSITIVE: 0,
            ClassificationType.DEVIATION: 0,
            ClassificationType.FIX_REQUIRED: 0,
            ClassificationType.UNDETERMINED: 0,
        }

        for result in results:
            counts[result.classification] += 1

        # タイトルを書き込む
        ws["A1"] = "分類結果サマリー"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:C1")

        # タイムスタンプを書き込む
        from datetime import datetime
        ws["A2"] = f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells("A2:C2")

        # 統計テーブルを書き込む
        headers = ["分類", "件数", "割合"]
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i)
            cell.value = header
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        row = 5
        for classification_type, count in counts.items():
            # 分類名
            cell_type = ws.cell(row=row, column=1)
            cell_type.value = classification_type.value
            cell_type.fill = PatternFill(
                start_color=self.CLASSIFICATION_COLORS[classification_type],
                end_color=self.CLASSIFICATION_COLORS[classification_type],
                fill_type="solid"
            )
            cell_type.border = thin_border

            # 件数
            cell_count = ws.cell(row=row, column=2)
            cell_count.value = count
            cell_count.alignment = Alignment(horizontal="right")
            cell_count.border = thin_border

            # 割合
            cell_pct = ws.cell(row=row, column=3)
            cell_pct.value = f"{count / total * 100:.1f}%" if total > 0 else "0%"
            cell_pct.alignment = Alignment(horizontal="right")
            cell_pct.border = thin_border

            row += 1

        # 合計行
        cell_total_label = ws.cell(row=row, column=1)
        cell_total_label.value = "合計"
        cell_total_label.font = Font(bold=True)
        cell_total_label.border = thin_border

        cell_total_count = ws.cell(row=row, column=2)
        cell_total_count.value = total
        cell_total_count.font = Font(bold=True)
        cell_total_count.alignment = Alignment(horizontal="right")
        cell_total_count.border = thin_border

        cell_total_pct = ws.cell(row=row, column=3)
        cell_total_pct.value = "100%"
        cell_total_pct.font = Font(bold=True)
        cell_total_pct.alignment = Alignment(horizontal="right")
        cell_total_pct.border = thin_border

        # 列幅を調整
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10

        wb.save(self.output_file)
        logger.info(f"Summary sheet added to {self.output_file}")
